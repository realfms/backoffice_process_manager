#!/usr/bin/python
# coding=utf-8

"""
Copyright 2012 Telefonica Investigación y Desarrollo, S.A.U

This file is part of Billing_PoC.

Billing_PoC is free software: you can redistribute it and/or modify it under the terms
of the GNU Affero General Public License as published by the Free Software Foundation, either
version 3 of the License, or (at your option) any later version.
Billing_PoC is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY; without even
the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU Affero
General Public License for more details.

You should have received a copy of the GNU Affero General Public License along with Billing_PoC.
If not, see http://www.gnu.org/licenses/.

For those usages not covered by the GNU Affero General Public License please contact with::mac@tid.es
"""

'''
Created on 20/06/2013

@author: mac@tid.es
'''


from openpyxl.reader.excel import load_workbook



wb = load_workbook(filename = r'templates/journal.xlsm', keep_vba=True)

wb.save(filename = 'results/journal.xlsm')



wb = load_workbook(filename = r'templates/revenue_report.xlsx', keep_vba=True)

ws = wb.get_active_sheet()

ws.append({0 : 'hola', 2 : 'adios'})

wb.save(filename = r'results/revenue_report.xlsx')