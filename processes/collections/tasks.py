#!/usr/bin/python
# coding=utf-8 

"""
Copyright 2012 Telefonica Investigación y Desarrollo, S.A.U

This file is part of Billing_PoC.

Billing_PoC is free software: you can redistribute it and/or modify it under the terms 
of the GNU Affero General Public License as published by the Free Software Foundation, either 
version 3 of the License, or (at your option) any later version.
Billing_PoC is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY; without even 
the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU Affero 
General Public License for more details.

You should have received a copy of the GNU Affero General Public License along with Billing_PoC. 
If not, see http://www.gnu.org/licenses/.

For those usages not covered by the GNU Affero General Public License please contact with::mac@tid.es
""" 

'''
Created on 15/10/2012

@author: mac@tid.es
'''

from celery import task

from processes.task_manager import TaskManager

from openpyxl.reader.excel import load_workbook

@task(ignore_result=True)
def generate_journal_task(success, orders, sp_id):
    tm = TaskManager()

    json = tm.get_subprocess_data(sp_id)

    return tm.process_task(sp_id, 'GENERATING JOURNAL', success, lambda : generate_journal(json, orders))

@task(ignore_result=True)
def generate_revenue_report_task(success, orders, sp_id):
    tm = TaskManager()
    return tm.process_task(sp_id, 'GENERATING REVENUES REPORT', success, lambda : generate_revenue_report(orders))

def generate_journal(json, orders):
    return (json, None)

def generate_revenue_report(orders):
    # wb = load_workbook(filename = r'processes/collections/templates/revenue_report.xlsx', keep_vba=True)
    wb = load_workbook(filename = 'processes/collections/templates/revenue_report.xlsx')

    ws = wb.get_active_sheet()

    [ws.append(order) for order in orders]

    filename = 'processes/collections/results/revenue_report.xlsx'

    wb.save(filename=filename)

    return ({'file_name': filename}, None)